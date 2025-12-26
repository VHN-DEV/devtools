#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Location Data Seeder Tool
Tool for fetching and exporting geographic data (countries, provinces, districts)
for database seeding in Botble CMS
"""

import os
import sys
import json
import csv
import time
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

try:
    import requests
except ImportError:
    print("[ERROR] Missing requests library. Install: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    print("[WARNING] Missing openpyxl library (for Excel export). Install: pip install openpyxl")
    EXCEL_AVAILABLE = False

# Import internal modules
from doc import get_documentation, show_help

# Constants
COUNTRIES_API = "https://restcountries.com/v3.1/all"
DEFAULT_COUNTRY_FIELDS = [
    'id', 'name', 'code', 'flag', 'capital', 'region', 'subregion',
    'population', 'area', 'languages', 'currencies', 'timezones'
]

DEFAULT_PROVINCE_FIELDS = [
    'id', 'name', 'code', 'region', 'area', 'population', 'country_id'
]

DEFAULT_DISTRICT_FIELDS = [
    'id', 'name', 'code', 'province_id', 'area', 'population', 'type'
]

DEFAULT_WARD_FIELDS = [
    'id', 'name', 'code', 'province_id', 'type'
]


def print_header():
    """Print tool header"""
    print("=" * 70)
    print("  [LOCATION DATA SEEDER]")
    print("  Tool for exporting geographic data for Botble CMS")
    print("=" * 70)
    print()


def format_size(size_bytes: int) -> str:
    """Format file size for readability"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} TB"


def get_downloads_path() -> str:
    """
    Get user's Downloads folder path

    Returns:
        str: Path to Downloads folder
    """
    try:
        # Windows
        if os.name == 'nt':
            # Thử dùng biến môi trường USERPROFILE
            user_profile = os.environ.get('USERPROFILE')
            if user_profile:
                downloads_path = os.path.join(user_profile, 'Downloads')
                if os.path.exists(downloads_path):
                    return downloads_path

            # Thử dùng biến môi trường HOMEPATH
            home_path = os.environ.get('HOMEPATH')
            if home_path:
                downloads_path = os.path.join(home_path, 'Downloads')
                if os.path.exists(downloads_path):
                    return downloads_path

        # Unix-like systems (Linux, macOS)
        else:
            home_dir = os.path.expanduser('~')
            downloads_path = os.path.join(home_dir, 'Downloads')
            if os.path.exists(downloads_path):
                return downloads_path

            # Thử một số đường dẫn khác trên Linux
            for alt_path in ['~/downloads', '~/Download']:
                alt_full_path = os.path.expanduser(alt_path)
                if os.path.exists(alt_full_path):
                    return alt_full_path

        # Fallback: sử dụng thư mục hiện tại
        print("[WARNING] Downloads folder not found, using current directory")
        return os.getcwd()

    except Exception as e:
        print(f"[WARN]  Lỗi khi lấy đường dẫn Downloads: {e}")
        return os.getcwd()


def load_json_data(file_path: str) -> list:
    """
    Load data from JSON file

    Args:
        file_path: Path to JSON file

    Returns:
        list: Data list from JSON file, or [] if error occurs
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            else:
                print(f"[WARN] File {file_path} does not contain valid JSON array")
                return []
    except FileNotFoundError:
        print(f"[WARN] Config file not found: {file_path}")
        return []
    except json.JSONDecodeError as e:
        print(f"[WARN] JSON parse error in file {file_path}: {e}")
        return []
    except Exception as e:
        print(f"[WARN] Error reading file {file_path}: {e}")
        return []


class LocationDataSeeder:
    """Main class for processing geographic data"""

    def __init__(self):
        self.countries = []
        self.provinces = []
        self.provinces_new = []  # 34 tỉnh thành mới
        self.districts = []
        self.wards = []  # Xã phường
        self.country_fields = DEFAULT_COUNTRY_FIELDS.copy()
        self.province_fields = DEFAULT_PROVINCE_FIELDS.copy()
        self.province_new_fields = DEFAULT_PROVINCE_FIELDS.copy()
        self.district_fields = DEFAULT_DISTRICT_FIELDS.copy()
        self.ward_fields = DEFAULT_WARD_FIELDS.copy()

    def fetch_countries_from_api(self) -> Tuple[bool, str]:
        """
        Fetch countries data from restcountries.com API

        Returns:
            tuple: (success, message)
        """
        try:
            print("[WORLD] Loading countries data from API...")
            response = requests.get(COUNTRIES_API, timeout=30)
            response.raise_for_status()

            countries_data = response.json()
            self.countries = []

            for idx, country in enumerate(countries_data, 1):
                # Map API data to our structure
                country_info = {
                    'id': idx,
                    'name': country.get('name', {}).get('common', ''),
                    'code': country.get('cca2', ''),
                    'flag': country.get('flag', ''),
                    'capital': country.get('capital', [None])[0] if country.get('capital') else None,
                    'region': country.get('region', ''),
                    'subregion': country.get('subregion', ''),
                    'population': country.get('population', 0),
                    'area': country.get('area', 0),
                    'languages': list(country.get('languages', {}).values()) if country.get('languages') else [],
                    'currencies': list(country.get('currencies', {}).keys()) if country.get('currencies') else [],
                    'timezones': country.get('timezones', []),
                    'latlng': country.get('latlng', []),
                    'borders': country.get('borders', [])
                }
                self.countries.append(country_info)

            return True, f"[SUCCESS] Successfully loaded {len(self.countries)} countries"

        except requests.RequestException as e:
            return False, f"[ERROR] API connection error: {str(e)}"
        except Exception as e:
            return False, f"[ERROR] Data processing error: {str(e)}"

    def get_vietnam_provinces(self) -> Tuple[bool, str]:
        """
        Load Vietnam provinces data from config file

        Returns:
            tuple: (success, message)
        """
        try:
            print("[VN] Loading Vietnam provinces data...")

            config_file = os.path.join(os.path.dirname(__file__), 'config', 'vietnam_provinces.json')
            vietnam_provinces = load_json_data(config_file)

            if not vietnam_provinces:
                return False, "[ERROR] Cannot load provinces data from config file"

            # Add country_id for Vietnam (assuming Vietnam has id 1 from countries data)
            for province in vietnam_provinces:
                province['country_id'] = 1  # Vietnam

            self.provinces = vietnam_provinces
            return True, f"[SUCCESS] Successfully loaded {len(self.provinces)} Vietnam provinces from config"

        except Exception as e:
            return False, f"[ERROR] Error loading provinces data: {str(e)}"

    def get_vietnam_districts(self) -> Tuple[bool, str]:
        """
        Load Vietnam districts data from config file

        Returns:
            tuple: (success, message)
        """
        try:
            print("[DISTRICT] Loading Vietnam districts data...")

            config_file = os.path.join(os.path.dirname(__file__), 'config', 'vietnam_districts.json')
            vietnam_districts = load_json_data(config_file)

            if not vietnam_districts:
                return False, "[ERROR] Cannot load districts data from config file"

            self.districts = vietnam_districts
            return True, f"[SUCCESS] Successfully loaded {len(self.districts)} districts from config"

        except Exception as e:
            return False, f"[ERROR] Error loading districts data: {str(e)}"

    def get_vietnam_provinces_new(self) -> Tuple[bool, str]:
        """
        Load new Vietnam provinces (34) data from config file

        Returns:
            tuple: (success, message)
        """
        try:
            print("[VN] Loading new Vietnam provinces (34) data...")

            config_file = os.path.join(os.path.dirname(__file__), 'config', 'vietnam_provinces_new.json')
            vietnam_provinces_new = load_json_data(config_file)

            if not vietnam_provinces_new:
                return False, "[ERROR] Cannot load new provinces data from config file"

            # Add country_id for Vietnam (assuming Vietnam has id 1 from countries data)
            for province in vietnam_provinces_new:
                province['country_id'] = 1  # Vietnam

            self.provinces_new = vietnam_provinces_new
            return True, f"[SUCCESS] Successfully loaded {len(self.provinces_new)} new Vietnam provinces from config (34 administrative units)"

        except Exception as e:
            return False, f"[ERROR] Error loading new provinces data: {str(e)}"

    def get_vietnam_wards(self) -> Tuple[bool, str]:
        """
        Load Vietnam wards data from config file

        Returns:
            tuple: (success, message)
        """
        try:
            print("[WARD] Loading Vietnam wards data...")

            config_file = os.path.join(os.path.dirname(__file__), 'config', 'vietnam_wards.json')
            vietnam_wards = load_json_data(config_file)

            if not vietnam_wards:
                return False, "[ERROR] Cannot load wards data from config file"

            self.wards = vietnam_wards
            return True, f"[SUCCESS] Successfully loaded {len(self.wards)} wards from config"

        except Exception as e:
            return False, f"[ERROR] Error loading wards data: {str(e)}"

    def get_vietnam_wards_new(self) -> Tuple[bool, str]:
        """
        Lấy dữ liệu xã phường Việt Nam mới từ file config

        Returns:
            tuple: (success, message)
        """
        try:
            print("[WARD]  Đang tải dữ liệu xã phường Việt Nam mới...")

            config_file = os.path.join(os.path.dirname(__file__), 'config', 'vietnam_wards_new.json')
            print(f"   [PATH] File config: {config_file}")
            vietnam_wards_new = load_json_data(config_file)

            if not vietnam_wards_new:
                return False, f"[ERROR] Không thể tải dữ liệu xã phường mới từ file config: {config_file}"

            self.wards = vietnam_wards_new
            return True, f"[SUCCESS] Đã tải thành công {len(self.wards)} xã phường mới từ config"

        except Exception as e:
            return False, f"[ERROR] Lỗi tải dữ liệu xã phường mới: {str(e)}"

    def customize_fields(self, data_type: str):
        """
        Cho phép user tùy chỉnh fields xuất

        Args:
            data_type: 'countries', 'provinces', 'provinces_new', 'districts', 'wards'
        """
        print(f"\n[CONFIG]  TÙY CHỈNH FIELDS XUẤT CHO {data_type.upper()}")

        if data_type == 'countries':
            current_fields = self.country_fields
            available_fields = [
                'id', 'name', 'code', 'flag', 'capital', 'region', 'subregion',
                'population', 'area', 'languages', 'currencies', 'timezones',
                'latlng', 'borders'
            ]
        elif data_type == 'provinces':
            current_fields = self.province_fields
            available_fields = ['id', 'name', 'code', 'region', 'area', 'population', 'country_id', 'type']
        elif data_type == 'provinces_new':
            current_fields = self.province_new_fields
            available_fields = ['id', 'name', 'code', 'region', 'area', 'population', 'country_id', 'type']
        elif data_type == 'districts':
            current_fields = self.district_fields
            available_fields = ['id', 'name', 'code', 'province_id', 'area', 'population', 'type']
        elif data_type == 'wards':
            current_fields = self.ward_fields
            available_fields = ['id', 'name', 'code', 'district_id', 'type']
        else:
            print("[ERROR] Loại dữ liệu không hợp lệ!")
            return

        print(f"\nFields hiện tại: {', '.join(current_fields)}")
        print(f"\nFields có sẵn: {', '.join(available_fields)}")

        choice = input("\nBạn muốn (1)Thêm fields, (2)Xóa fields, (3)Reset mặc định, (4)Giữ nguyên: ").strip()

        if choice == '1':
            # Thêm fields
            print("\nNhập fields muốn thêm (cách nhau bởi dấu phẩy):")
            new_fields_input = input("> ").strip()
            new_fields = [f.strip() for f in new_fields_input.split(',') if f.strip()]

            for field in new_fields:
                if field in available_fields and field not in current_fields:
                    current_fields.append(field)
                    print(f"[SUCCESS] Đã thêm field: {field}")
                elif field not in available_fields:
                    print(f"[WARN]  Field '{field}' không có sẵn, bỏ qua")
                else:
                    print(f"[WARN]  Field '{field}' đã tồn tại")

        elif choice == '2':
            # Xóa fields
            print("\nNhập fields muốn xóa (cách nhau bởi dấu phẩy):")
            remove_fields_input = input("> ").strip()
            remove_fields = [f.strip() for f in remove_fields_input.split(',') if f.strip()]

            for field in remove_fields:
                if field in current_fields:
                    current_fields.remove(field)
                    print(f"[SUCCESS] Đã xóa field: {field}")
                else:
                    print(f"[WARN]  Field '{field}' không tồn tại trong danh sách hiện tại")

        elif choice == '3':
            # Reset mặc định
            if data_type == 'countries':
                self.country_fields = DEFAULT_COUNTRY_FIELDS.copy()
            elif data_type == 'provinces':
                self.province_fields = DEFAULT_PROVINCE_FIELDS.copy()
            elif data_type == 'districts':
                self.district_fields = DEFAULT_DISTRICT_FIELDS.copy()
            print("[SUCCESS] Đã reset về fields mặc định")

        elif choice == '4':
            print("[SUCCESS] Giữ nguyên fields hiện tại")

        else:
            print("[ERROR] Lựa chọn không hợp lệ!")

        print(f"\nFields cuối cùng: {', '.join(current_fields)}")

    def filter_data_by_fields(self, data: List[Dict], fields: List[str]) -> List[Dict]:
        """
        Lọc dữ liệu theo fields được chọn

        Args:
            data: Danh sách dữ liệu gốc
            fields: Danh sách fields muốn giữ lại

        Returns:
            Danh sách dữ liệu đã lọc
        """
        filtered_data = []
        for item in data:
            filtered_item = {}
            for field in fields:
                if field in item:
                    filtered_item[field] = item[field]
            filtered_data.append(filtered_item)
        return filtered_data

    def export_json(self, output_path: str) -> Tuple[bool, str]:
        """
        Xuất dữ liệu ra file JSON

        Args:
            output_path: Đường dẫn file xuất

        Returns:
            tuple: (success, message)
        """
        try:
            print("[DATA] Đang xuất dữ liệu JSON...")

            # Filter data theo fields đã chọn
            countries_data = self.filter_data_by_fields(self.countries, self.country_fields)
            provinces_data = self.filter_data_by_fields(self.provinces, self.province_fields)
            provinces_new_data = self.filter_data_by_fields(self.provinces_new, self.province_new_fields)
            districts_data = self.filter_data_by_fields(self.districts, self.district_fields)
            wards_data = self.filter_data_by_fields(self.wards, self.ward_fields)

            # Tạo structure JSON
            export_data = {
                'metadata': {
                    'exported_at': time.strftime('%Y-%m-%d %H:%M:%S'),
                    'countries_count': len(countries_data),
                    'provinces_count': len(provinces_data),
                    'provinces_new_count': len(provinces_new_data),
                    'districts_count': len(districts_data),
                    'wards_count': len(wards_data),
                    'fields': {
                        'countries': self.country_fields,
                        'provinces': self.province_fields,
                        'provinces_new': self.province_new_fields,
                        'districts': self.district_fields,
                        'wards': self.ward_fields
                    }
                },
                'countries': countries_data,
                'provinces': provinces_data,
                'provinces_new': provinces_new_data,
                'districts': districts_data,
                'wards': wards_data
            }

            # Ghi file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)

            file_size = os.path.getsize(output_path)
            return True, f"[SUCCESS] Xuất JSON thành công!\n   [PATH] Đường dẫn: {output_path}\n   [DATA] Kích thước: {format_size(file_size)}"

        except Exception as e:
            return False, f"[ERROR] Lỗi xuất JSON: {str(e)}"

    def export_excel(self, output_path: str) -> Tuple[bool, str]:
        """
        Xuất dữ liệu ra file Excel

        Args:
            output_path: Đường dẫn file xuất

        Returns:
            tuple: (success, message)
        """
        if not EXCEL_AVAILABLE:
            return False, "[ERROR] Thư viện openpyxl không có sẵn. Cài đặt: pip install openpyxl"

        try:
            print("📈 Đang xuất dữ liệu Excel...")

            # Filter data theo fields đã chọn
            countries_data = self.filter_data_by_fields(self.countries, self.country_fields)
            provinces_data = self.filter_data_by_fields(self.provinces, self.province_fields)
            provinces_new_data = self.filter_data_by_fields(self.provinces_new, self.province_new_fields)
            districts_data = self.filter_data_by_fields(self.districts, self.district_fields)
            wards_data = self.filter_data_by_fields(self.wards, self.ward_fields)

            # Tạo workbook
            wb = Workbook()

            # Sheet Countries
            ws_countries = wb.active
            ws_countries.title = "Countries"
            ws_countries.sheet_properties.tabColor = "FF0000"  # Red

            # Header cho Countries
            for col, field in enumerate(self.country_fields, 1):
                cell = ws_countries.cell(row=1, column=col, value=field.title())
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

            # Data cho Countries
            for row, country in enumerate(countries_data, 2):
                for col, field in enumerate(self.country_fields, 1):
                    value = country.get(field, '')
                    # Convert lists to strings for Excel
                    if isinstance(value, list):
                        value = ', '.join(str(v) for v in value)
                    ws_countries.cell(row=row, column=col, value=value)

            # Sheet Provinces
            ws_provinces = wb.create_sheet("Provinces")
            ws_provinces.sheet_properties.tabColor = "00FF00"  # Green

            # Header cho Provinces
            for col, field in enumerate(self.province_fields, 1):
                cell = ws_provinces.cell(row=1, column=col, value=field.title())
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

            # Data cho Provinces
            for row, province in enumerate(provinces_data, 2):
                for col, field in enumerate(self.province_fields, 1):
                    value = province.get(field, '')
                    ws_provinces.cell(row=row, column=col, value=value)

            # Sheet Districts
            ws_districts = wb.create_sheet("Districts")
            ws_districts.sheet_properties.tabColor = "0000FF"  # Blue

            # Header cho Districts
            for col, field in enumerate(self.district_fields, 1):
                cell = ws_districts.cell(row=1, column=col, value=field.title())
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6FF", end_color="E6E6FF", fill_type="solid")

            # Data cho Districts
            for row, district in enumerate(districts_data, 2):
                for col, field in enumerate(self.district_fields, 1):
                    value = district.get(field, '')
                    ws_districts.cell(row=row, column=col, value=value)

            # Sheet Provinces New
            if provinces_new_data:
                ws_provinces_new = wb.create_sheet("Provinces_New")
                ws_provinces_new.sheet_properties.tabColor = "FFFF00"  # Yellow

                # Header cho Provinces New
                for col, field in enumerate(self.province_new_fields, 1):
                    cell = ws_provinces_new.cell(row=1, column=col, value=field.title())
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid")

                # Data cho Provinces New
                for row, province in enumerate(provinces_new_data, 2):
                    for col, field in enumerate(self.province_new_fields, 1):
                        value = province.get(field, '')
                        ws_provinces_new.cell(row=row, column=col, value=value)

            # Sheet Wards
            if wards_data:
                ws_wards = wb.create_sheet("Wards")
                ws_wards.sheet_properties.tabColor = "FF00FF"  # Magenta

                # Header cho Wards
                for col, field in enumerate(self.ward_fields, 1):
                    cell = ws_wards.cell(row=1, column=col, value=field.title())
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFE6FF", end_color="FFE6FF", fill_type="solid")

                # Data cho Wards
                for row, ward in enumerate(wards_data, 2):
                    for col, field in enumerate(self.ward_fields, 1):
                        value = ward.get(field, '')
                        ws_wards.cell(row=row, column=col, value=value)

            # Lưu file
            wb.save(output_path)

            file_size = os.path.getsize(output_path)
            return True, f"[SUCCESS] Xuất Excel thành công!\n   [PATH] Đường dẫn: {output_path}\n   [DATA] Kích thước: {format_size(file_size)}"

        except Exception as e:
            return False, f"[ERROR] Lỗi xuất Excel: {str(e)}"

    def export_sql(self, output_path: str, table_prefix: str = '') -> Tuple[bool, str]:
        """
        Xuất dữ liệu ra file SQL seed

        Args:
            output_path: Đường dẫn file xuất
            table_prefix: Prefix cho tên bảng (vd: 'bc_')

        Returns:
            tuple: (success, message)
        """
        try:
            print("[SQL]  Đang xuất dữ liệu SQL...")

            sql_content = []
            sql_content.append("-- Location Data Seed File")
            sql_content.append(f"-- Generated at: {time.strftime('%Y-%m-%d %H:%M:%S')}")
            sql_content.append("-- For Botble CMS Database")
            sql_content.append("")

            # Filter data theo fields đã chọn
            countries_data = self.filter_data_by_fields(self.countries, self.country_fields)
            provinces_data = self.filter_data_by_fields(self.provinces, self.province_fields)
            provinces_new_data = self.filter_data_by_fields(self.provinces_new, self.province_new_fields)
            districts_data = self.filter_data_by_fields(self.districts, self.district_fields)
            wards_data = self.filter_data_by_fields(self.wards, self.ward_fields)

            # Countries table
            if countries_data:
                sql_content.append(f"-- {table_prefix}countries table")
                for country in countries_data:
                    values = []
                    for field in self.country_fields:
                        value = country.get(field, '')
                        if isinstance(value, str):
                            values.append(f"'{value.replace(chr(39), chr(39) + chr(39))}'")  # Escape single quotes
                        elif isinstance(value, list):
                            values.append(f"'{json.dumps(value, ensure_ascii=False)}'")
                        elif value is None:
                            values.append('NULL')
                        else:
                            values.append(str(value))

                    sql_content.append(f"INSERT INTO {table_prefix}countries ({', '.join(self.country_fields)}) VALUES ({', '.join(values)});")

                sql_content.append("")

            # Provinces table
            if provinces_data:
                sql_content.append(f"-- {table_prefix}provinces table")
                for province in provinces_data:
                    values = []
                    for field in self.province_fields:
                        value = province.get(field, '')
                        if isinstance(value, str):
                            values.append(f"'{value.replace(chr(39), chr(39) + chr(39))}'")  # Escape single quotes
                        elif isinstance(value, list):
                            values.append(f"'{json.dumps(value, ensure_ascii=False)}'")
                        elif value is None:
                            values.append('NULL')
                        else:
                            values.append(str(value))

                    sql_content.append(f"INSERT INTO {table_prefix}provinces ({', '.join(self.province_fields)}) VALUES ({', '.join(values)});")

                sql_content.append("")

            # Districts table
            if districts_data:
                sql_content.append(f"-- {table_prefix}districts table")
                for district in districts_data:
                    values = []
                    for field in self.district_fields:
                        value = district.get(field, '')
                        if isinstance(value, str):
                            values.append(f"'{value.replace(chr(39), chr(39) + chr(39))}'")  # Escape single quotes
                        elif isinstance(value, list):
                            values.append(f"'{json.dumps(value, ensure_ascii=False)}'")
                        elif value is None:
                            values.append('NULL')
                        else:
                            values.append(str(value))

                    sql_content.append(f"INSERT INTO {table_prefix}districts ({', '.join(self.district_fields)}) VALUES ({', '.join(values)});")

                sql_content.append("")

            # Provinces New table
            if provinces_new_data:
                sql_content.append(f"-- {table_prefix}provinces_new table")
                for province in provinces_new_data:
                    values = []
                    for field in self.province_new_fields:
                        value = province.get(field, '')
                        if isinstance(value, str):
                            values.append(f"'{value.replace(chr(39), chr(39) + chr(39))}'")  # Escape single quotes
                        elif isinstance(value, list):
                            values.append(f"'{json.dumps(value, ensure_ascii=False)}'")
                        elif value is None:
                            values.append('NULL')
                        else:
                            values.append(str(value))

                    sql_content.append(f"INSERT INTO {table_prefix}provinces_new ({', '.join(self.province_new_fields)}) VALUES ({', '.join(values)});")

                sql_content.append("")

            # Wards table
            if wards_data:
                sql_content.append(f"-- {table_prefix}wards table")
                for ward in wards_data:
                    values = []
                    for field in self.ward_fields:
                        value = ward.get(field, '')
                        if isinstance(value, str):
                            values.append(f"'{value.replace(chr(39), chr(39) + chr(39))}'")  # Escape single quotes
                        elif isinstance(value, list):
                            values.append(f"'{json.dumps(value, ensure_ascii=False)}'")
                        elif value is None:
                            values.append('NULL')
                        else:
                            values.append(str(value))

                    sql_content.append(f"INSERT INTO {table_prefix}wards ({', '.join(self.ward_fields)}) VALUES ({', '.join(values)});")

            # Ghi file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(sql_content))

            file_size = os.path.getsize(output_path)
            return True, f"[SUCCESS] Xuất SQL thành công!\n   [PATH] Đường dẫn: {output_path}\n   [DATA] Kích thước: {format_size(file_size)}"

        except Exception as e:
            return False, f"[ERROR] Lỗi xuất SQL: {str(e)}"

    def show_data_summary(self):
        """Hiển thị tổng quan dữ liệu"""
        print("\n[DATA] TỔNG QUAN DỮ LIỆU")
        print("=" * 50)

        if self.countries:
            print(f"[WORLD] Countries: {len(self.countries)} quốc gia")
            print(f"   Fields: {', '.join(self.country_fields)}")

        if self.provinces:
            print(f"[VN] Provinces (63): {len(self.provinces)} tỉnh thành")
            print(f"   Fields: {', '.join(self.province_fields)}")

        if self.provinces_new:
            print(f"[VN] Provinces New (34): {len(self.provinces_new)} tỉnh thành mới")
            print(f"   Fields: {', '.join(self.province_new_fields)}")

        if self.districts:
            print(f"[DISTRICT]  Districts: {len(self.districts)} quận huyện")
            print(f"   Fields: {', '.join(self.district_fields)}")

        if self.wards:
            print(f"[WARD]  Wards: {len(self.wards)} xã phường")
            print(f"   Fields: {', '.join(self.ward_fields)}")
            print(f"   Debug: wards type = {type(self.wards)}, len = {len(self.wards)}")
        else:
            print(f"[WARD]  Wards: 0 xã phường (empty)")

        total_records = len(self.countries) + len(self.provinces) + len(self.provinces_new) + len(self.districts) + len(self.wards)
        if total_records > 0:
            print(f"\n📈 Tổng cộng: {total_records} bản ghi")
        else:
            print("[ERROR] Chưa có dữ liệu nào được tải!")

        print("=" * 50)


def main():
    """Hàm chính - Menu chính của tool"""
    print_header()

    seeder = LocationDataSeeder()

    while True:
        print("\n===== MAIN MENU =====")
        print("1. [DOWNLOAD] Load countries data")
        print("2. [DOWNLOAD] Load Vietnam provinces (63)")
        print("3. [DOWNLOAD] Load Vietnam provinces new (34)")
        print("4. [DOWNLOAD] Load Vietnam districts")
        print("5. [DOWNLOAD] Load Vietnam wards (63 provinces)")
        print("6. [DOWNLOAD] Load Vietnam wards new (34 provinces - 2-level system)")
        print("7. [CONFIG] Customize export fields")
        print("8. [VIEW] View data summary")
        print("9. [EXPORT] Export to JSON")
        print("10. [EXPORT] Export to Excel")
        print("11. [EXPORT] Export to SQL")
        print("12. [HELP] Help & Usage")
        print("0. [EXIT] Exit")

        choice = input("\nChọn chức năng (0-12): ").strip()

        if choice == "0":
            print("\n👋 Cảm ơn đã sử dụng Location Data Seeder!")
            break

        elif choice == "1":
            # Tải dữ liệu quốc gia
            success, message = seeder.fetch_countries_from_api()
            print(f"\n{message}")

        elif choice == "2":
            # Tải dữ liệu tỉnh thành VN (63)
            success, message = seeder.get_vietnam_provinces()
            print(f"\n{message}")

        elif choice == "3":
            # Tải dữ liệu tỉnh thành VN mới (34)
            success, message = seeder.get_vietnam_provinces_new()
            print(f"\n{message}")

        elif choice == "4":
            # Tải dữ liệu quận huyện VN
            success, message = seeder.get_vietnam_districts()
            print(f"\n{message}")

        elif choice == "5":
            # Tải dữ liệu xã phường VN (63)
            success, message = seeder.get_vietnam_wards()
            print(f"\n{message}")

        elif choice == "6":
            # Tải dữ liệu xã phường VN mới (34)
            success, message = seeder.get_vietnam_wards_new()
            print(f"\n{message}")

        elif choice == "6":
            # Tùy chỉnh fields
            print("\nChọn loại dữ liệu để tùy chỉnh:")
            print("1. Countries (quốc gia)")
            print("2. Provinces (tỉnh thành)")
            print("3. Provinces New (tỉnh thành mới)")
            print("4. Districts (quận huyện)")
            print("5. Wards (xã phường)")

            field_choice = input("\nChọn (1-5): ").strip()

            if field_choice == "1":
                seeder.customize_fields('countries')
            elif field_choice == "2":
                seeder.customize_fields('provinces')
            elif field_choice == "3":
                seeder.customize_fields('provinces_new')
            elif field_choice == "4":
                seeder.customize_fields('districts')
            elif field_choice == "5":
                seeder.customize_fields('wards')
            else:
                print("[ERROR] Lựa chọn không hợp lệ!")

        elif choice == "7":
            # Xem tổng quan
            seeder.show_data_summary()

        elif choice == "8":
            # Xuất JSON
            if not any([seeder.countries, seeder.provinces, seeder.provinces_new, seeder.districts, seeder.wards]):
                print("\n[ERROR] Chưa có dữ liệu nào để xuất! Hãy tải dữ liệu trước.")
                continue

            downloads_path = get_downloads_path()
            default_path = os.path.join(downloads_path, f"location_data_{time.strftime('%Y%m%d_%H%M%S')}.json")
            print(f"[PATH] Đường dẫn mặc định: {default_path}")
            output_file = input("Nhập tên file xuất (Enter để dùng mặc định): ").strip()
            if not output_file:
                output_file = default_path

            success, message = seeder.export_json(output_file)
            print(f"\n{message}")

        elif choice == "9":
            # Xuất Excel
            if not EXCEL_AVAILABLE:
                print("\n[ERROR] Không thể xuất Excel vì thiếu thư viện openpyxl!")
                print("Cài đặt bằng: pip install openpyxl")
                continue

            if not any([seeder.countries, seeder.provinces, seeder.provinces_new, seeder.districts, seeder.wards]):
                print("\n[ERROR] Chưa có dữ liệu nào để xuất! Hãy tải dữ liệu trước.")
                continue

            downloads_path = get_downloads_path()
            default_path = os.path.join(downloads_path, f"location_data_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
            print(f"[PATH] Đường dẫn mặc định: {default_path}")
            output_file = input("Nhập tên file xuất (Enter để dùng mặc định): ").strip()
            if not output_file:
                output_file = default_path

            success, message = seeder.export_excel(output_file)
            print(f"\n{message}")

        elif choice == "10":
            # Xuất SQL
            if not any([seeder.countries, seeder.provinces, seeder.provinces_new, seeder.districts, seeder.wards]):
                print("\n[ERROR] Chưa có dữ liệu nào để xuất! Hãy tải dữ liệu trước.")
                continue

            downloads_path = get_downloads_path()
            default_path = os.path.join(downloads_path, f"location_seed_{time.strftime('%Y%m%d_%H%M%S')}.sql")
            print(f"[PATH] Đường dẫn mặc định: {default_path}")
            output_file = input("Nhập tên file xuất (Enter để dùng mặc định): ").strip()
            if not output_file:
                output_file = default_path

            table_prefix = input("Prefix cho tên bảng (vd: bc_, Enter để bỏ trống): ").strip()

            success, message = seeder.export_sql(output_file, table_prefix)
            print(f"\n{message}")

        elif choice == "11":
            # Hướng dẫn
            show_help()

        else:
            print("[ERROR] Lựa chọn không hợp lệ! Vui lòng chọn lại.")

        input("\nNhấn Enter để tiếp tục...")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n[ERROR] Đã hủy!")
    except Exception as e:
        print(f"\n[ERROR] Lỗi: {e}")
